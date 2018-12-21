# -*- coding: utf-8 -*-
"""
Created on Thu Jan 25 10:05:45 2018

@author: Maxwell Ledermann
For questions or assistance, email maxwell.ledermann@gmail.com
"""

import cv2
import numpy as np
import math
import datetime
import pickle
import openpyxl as pyxl
import keyboard
import sys
import os
import createcodes

print("3414 Hackbots POWER UP Scouting Program")
print("Press 's' to switch the active camera")
print("Press 'Esc' to exit the program\n")

codes_directory="Nonacodes" #Default is "Nonacodes", change if using a different code storage folder.
createcodes.createCodes() #Creates codes in "Nonacodes" folder. Irrelevant if codes_directory is changed

pickle.dump((0,0,False), open('entry data.p','wb'))
rescan=False
last_scan=[-1,-1] #-1 is used as a placeholder value to indicate no scan has been completed yet.
cv2.namedWindow('Camera Capture')
cv2.imshow("Scan",np.empty((480,640,3)))

#Select a capture device and start recording a live capture from a camera.
capture_device = 0
cap = cv2.VideoCapture(capture_device)
ret, __ = cap.read()
if ret == False: #If a capture fails to start no error is given, but trying to read an empty capture returns false.
    cap = cv2.VideoCapture(capture_device+1) #Occasionally a capture will fail to start despite a camera being connected. I have found the best way to remedy this is to
    cap = cv2.VideoCapture(capture_device) #change the capture device and then change it back.
    ret, __ = cap.read()
    if ret == False: #If the capture still cannot be started, make sure that the camera feed is not being used by another application.
        input("No camera detected. Please connect a compatible camera and retry.\nPress enter to exit the program")
        sys.exit()
        
#Does a main scouting spreadsheet exist, if not, create one.
try:
    wb = pyxl.load_workbook('scouting.xlsx')
    ws = wb.active
except:
    wb = pyxl.Workbook()
    ws = wb.active
    #Titles of columns in spreadsheet. These names do not have to exactly match the variables variable names.
    header = ["Time","Team Number","Match Number", "Alliance Station","Starting Position","Plate Config", \
    "Crossed Baseline","Preload Cube","Second Cube","Teleop Scale","Teleop Switch","Teleop Op Switch", \
    "Teleop Vault", "Climbed","Parked","Lifted One","Lifted Two","Was Lifted"]
    for header_column in header:
        ws.cell(row=1, column=(header.index(header_column)+1)).value = header_column

#From the list of coordinate points, returns which is closest to the target point
def find_corner(target,points):
    lowest_distance = 100000
    for point in points:
        if (math.sqrt((point[0]-target[0])**2 + (point[1]-target[1])**2)) < lowest_distance:
            coord = point
            lowest_distance = math.sqrt((point[0]-target[0])**2 + (point[1]-target[1])**2)
    return(coord)

#Returns the list of contours and the thresholded image from an input image
def extract_contours(image):
    image = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
    image = cv2.adaptiveThreshold(image,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,23,6) #Refer to readme.txt for threshold calibration instructions
    ret, contours, contour_hierarchy = cv2.findContours(image,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    return contours,image

def code_matching(code,variables):
    #What code number is associated with each variable change. These do not have to be in any particular order.
    false_positive = False
    try:
        if code_to_match == 1:
            variables["starting_position"].remove("left")
        elif code_to_match == 2:
            variables["starting_position"].remove("center")
        elif code_to_match == 3:
            variables["starting_position"].remove("right")
        elif code_to_match == 4:
            variables["crossed_baseline"] = False
        elif code_to_match == 5:
            variables["preload_cube"].remove("switch")
        elif code_to_match == 6:
            variables["preload_cube"].remove("scale")
        elif code_to_match == 7:
            variables["second_cube"].remove("switch")
        elif code_to_match == 8:
            variables["second_cube"].remove("scale")
        elif code_to_match == 9:
            variables["found_scale"] -= 1
        elif code_to_match == 10:
            variables["found_switch"] -= 1
        elif code_to_match == 11:
            variables["found_op_switch"] -= 1
        elif code_to_match == 12:
            variables["found_vault"] -= 1
        elif code_to_match == 13:
            variables["climbed"] = False
        elif code_to_match == 14:
            variables["parked"] = False
        elif code_to_match == 15:
            variables["lift_one"] = False
        elif code_to_match == 16:
            variables["lift_two"] = False
        elif code_to_match == 17:
            variables["was_lifted"] = False
        elif code_to_match == 18:
            variables["plate_config"].remove("LLL")
        elif code_to_match == 19:
            variables["plate_config"].remove("RRR")
        elif code_to_match == 20:
            variables["plate_config"].remove("LRL")
        elif code_to_match == 21:
            variables["plate_config"].remove("RLR")
        elif code_to_match == 79:
            variables["alliance_station"] = "Red1"
        elif code_to_match == 80:
            variables["alliance_station"] = "Red2"
        elif code_to_match == 81:
            variables["alliance_station"] = "Red3"
        elif code_to_match == 82:
            variables["alliance_station"] = "Blue1"
        elif code_to_match == 83:
            variables["alliance_station"] = "Blue2"
        elif code_to_match == 84:
            variables["alliance_station"] = "Blue3"
    except:
        print("BAD SCAN: Too many copies of code" + str(code_to_match) + " detected.")
        false_positive = True
    return variables, false_positive
     
while(True):
    #Press the s key to switch between successive cameras. Can switch between any number of connected devices. 
    if keyboard.is_pressed('s'):
        while keyboard.is_pressed('s'):
            pass
        capture_device += 1
        cap = cv2.VideoCapture(capture_device)
        ret, __ = cap.read()
        if ret == False:
            capture_device = 0
            cap = cv2.VideoCapture(capture_device)
            
    #Press the m key to add the team number and match number of the previous scan to a text document for manual review later.
    #Will only run if last_scan is not [-1,-1], which is only true if a scan has already been completed.
    if keyboard.is_pressed('m') and last_scan != [-1,-1]:
        while keyboard.is_pressed('m'):
            pass
        if os.path.isfile("Needs asjustment.txt"):
            with open("Needs asjustment.txt", "a") as myfile:
                myfile.write("\nMatch: " + str(last_scan[1]) + "      Team: " + str(last_scan[0]))
                print("Scan marked for manual data adjustment")
        else:
            with open("Needs asjustment.txt", "a") as myfile:
                myfile.write("Match: " + str(last_scan[1]) + "      Team: " + str(last_scan[0]))
                print("Scan marked for manual data adjustment")
        
    team_number, match_number, SCAN = pickle.load(open("entry data.p", "rb"))
    
    #Press the r key to initiate a scan using the last used team and match numbers.
    if keyboard.is_pressed('r') and last_scan != [-1,-1]:
        while keyboard.is_pressed('r'):
            pass
        rescan = True
        SCAN = True
        
    ret, frame = cap.read()
    cv2.imshow("Camera Capture",frame)
    cv2.waitKey(1)
    
    #As a debug tool, press ctrl and space together to trigger scanning with default values of 0 for match number and team number
    if keyboard.is_pressed('ctrl+space') == True:
        SCAN = True

    if SCAN == True:
        SCAN = False
        pickle.dump((0,0,SCAN), open('entry data.p','wb'))
        save_data = True
        clear = lambda: os.system('cls')
        clear()
        print("3414 Hackbots POWER UP Scouting Program")
        print("Press 's' to switch the active camera")
        print("Press 'Esc' to exit the program\n")
        print("Now scanning frame. Press 'c' to cancel the scan and not save any data.")
        
        #Variables should be in the order they are appear on the spreadsheet left to right.
        variables = {"alliance_station":"",
                        "plate_config":["LLL","RRR","LRL","RLR"],
                        "starting_position":["left","center","right"],
                        "crossed_baseline":True,
                        "preload_cube":["switch","scale"],
                        "second_cube":["switch","scale"],
                        "found_scale":10,
                        "found_switch":10,
                        "found_op_switch":10,
                        "found_vault":10,
                        "climbed":True,
                        "parked":True,
                        "lift_one":True,
                        "lift_two":True,
                        "was_lifted":True,}
        
        codes_highlighted = frame.copy()
        cv2.imshow("Scan",codes_highlighted)
        cv2.waitKey(1)
        
        contours,frame_threshold = extract_contours(codes_highlighted)
        
        #Only use contours with a perimeter length between a min and max value. Helps reduce the number of irellevant contours
        contours_limited = [i for k,i in enumerate(contours) if cv2.arcLength(contours[k],1)>100 and cv2.arcLength(contours[k],1)<250]
        
        for contour_num,current_contour in enumerate(contours_limited):
            print("Scanning " + str(round(float(float((contour_num)/float(len(contours_limited))*100.0)),1)) + "%")
            
            rect = cv2.minAreaRect(current_contour)
            box = cv2.boxPoints(rect)
            box = np.int32(box)
            
            contour_moments=cv2.moments(current_contour)
            center_x = int(contour_moments['m10']/contour_moments['m00'])
            center_y = int(contour_moments['m01']/contour_moments['m00'])            
            match_point_tl = [center_x-10000,center_y-10000]
            match_point_tr = [center_x+10000,center_y-10000]
            match_point_bl = [center_x-10000,center_y+10000]
            match_point_br = [center_x+10000,center_y+10000]
            
            points = current_contour[:,0,:]
            top_left = find_corner(match_point_tl,points)
            top_right = find_corner(match_point_tr,points)
            bottom_left = find_corner(match_point_bl,points)
            bottom_right = find_corner(match_point_br,points)
            warp_input_corners = np.float32([top_left,top_right,bottom_left,bottom_right])
            
            
            matching_scale=100 #Low values are innacurate, high values are slower, default 100
            output_corners = np.float32([[0,0],[matching_scale-1,0],[0,matching_scale-1],[matching_scale-1,matching_scale-1]]) #coordinates for the  
            transform_matrix = cv2.getPerspectiveTransform(warp_input_corners,output_corners) #3x3 transformation matrix for running a perspective transfrom function.
            object_warped = cv2.warpPerspective(frame_threshold,transform_matrix,(matching_scale-1,matching_scale-1))   
            
            for code_to_match in list(range(1,22)) + list(range(79,85)):  
                if keyboard.is_pressed('c'):
                    save_data = False
                    break
                
                match_template = cv2.imread(codes_directory + '/code' + str(code_to_match) + '.png',0)
                template_corners = np.float32([[0,0],[399,0],[0,399],[399,399]])
                transform_matrix = cv2.getPerspectiveTransform(template_corners,output_corners)
                match_template = cv2.warpPerspective(match_template,transform_matrix,(matching_scale-1,matching_scale-1))
                
                method = eval('cv2.TM_CCOEFF_NORMED')
                try:
                    res = cv2.matchTemplate(object_warped,match_template,method)
                except:
                    print("code" + str(code_to_match) + ".png not found in " + codes_directory + " directory. Please run 'Create Codes.py' and retry.")
                    input("Press enter with this console selected to exit")
                    cap.release()
                    cv2.destroyAllWindows()
                    sys.exit()
                
                min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
                if max_val > .7: #If you notice that occasionally a code won't get scanned when it should, decrease this by 0.01 or 0.02.
                                 #If you get BAD SCAN errors or codes being mistaken for eachother then increase.
                    false_positive = False
                    variables,false_positive = code_matching(code_to_match,variables)
                    if false_positive == False:
                        cv2.drawContours(codes_highlighted,[box],0,(0,255,0),2)
                    else:
                        cv2.drawContours(codes_highlighted,[box],0,(0,0,255),2)
                    cv2.imshow("Scan",codes_highlighted)
                    cv2.waitKey(1)
                    break
                
            if save_data == False:
                break

        if save_data == True:
            print("Scan Complete")

            font = cv2.FONT_HERSHEY_SIMPLEX
            cv2.putText(codes_highlighted,str("Team Number: " + str(team_number)),(5,15),font,.5,(0,0,0),3,cv2.LINE_AA)
            cv2.putText(codes_highlighted,str("Team Number: " + str(team_number)),(5,15),font,.5,(255,255,255),1,cv2.LINE_AA)
            cv2.putText(codes_highlighted,str("Match Number: " + str(match_number)),(5,30),font,.5,(0,0,0),3,cv2.LINE_AA)
            cv2.putText(codes_highlighted,str("Match Number: " +str( match_number)),(5,30),font,.5,(255,255,255),1,cv2.LINE_AA)
            cv2.imshow("Scan",codes_highlighted)
            
            export_data = list(variables.values())
            
            #Converts list type variables to strings. If the list length is 0 or greater than 1 then it is replaced with a error string.
            for data_index in range(len(export_data)):
                if isinstance(export_data[data_index],list):
                    if len(export_data[data_index]) == 1:
                        export_data[data_index] = str(export_data[data_index][0])
                    elif len(export_data[data_index]) == 0:
                        export_data[data_index] = "BAD SCAN"
                    elif len(export_data[data_index]) > 1:
                        export_data[data_index] = "BAD SCAN"
                else:
                    export_data[data_index] = str(export_data[data_index])
            
            #Adds time of data collection, the team number and match number as the first three columns to be uploaded to the spreadsheet.
            now = datetime.datetime.now()
            export_data.insert(0,now)
            export_data.insert(1,team_number)
            export_data.insert(2,match_number)
            ws.append(export_data)
            
            #If this scan was triggered by rescanning delete the last row in scouting.xlsx
            if rescan == True:
                last_filled_row = ws.max_row
                ws.delete_rows(last_filled_row,1)
                team_number = last_scan[0]
                match_number = last_scan[1]
                rescan = False
            
            if not os.path.exists('Spreadsheet backups'):
                os.mkdir('Spreadsheet backups')
            while(True):
                try:
                    wb.save('scouting.xlsx')
                    wb.save('Spreadsheet backups/scouting backup ' + str(match_number) + '.xlsx')
                    break
                except:
                    if input("Please close scouting.xlsx | Press enter when closed to continue| Type 'exit' to exit") == "exit":
                        cap.release()
                        cv2.destroyAllWindows()
                        sys.exit()
                        
            if not os.path.exists('Saved scans'):
                os.mkdir('Saved scans')
            save_name = "Saved scans/match" + str(match_number) + "-team" + str(team_number) + ".png"
            cv2.imwrite(save_name, codes_highlighted)
            print("Data Saved")
            print("Press 'm' to mark this scan for manual adjustment later")
            print("Press 'r' to rescan using the previous team and match numbers")
            last_scan = [team_number,match_number]
            
        elif save_data == False:
            rescan = False
            print("Scan canceled, no data saved \n")
            
    if keyboard.is_pressed('esc'):
        break

cap.release()
cv2.destroyAllWindows()